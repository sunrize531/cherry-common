from json import dumps, loads
from cherrycommon.dictutils import flatten_value
from openpyxl.workbook import Workbook as _Workbook
from openpyxl import style
from openpyxl import load_workbook


HEADER = 'header'
CELL = 'cell'

_CELL_FORMAT = style.Style()
_CELL_FORMAT.font.name = 'courier'

_HEADER_FORMAT = style.Style()
_HEADER_FORMAT.font.name = 'courier'
_HEADER_FORMAT.font.bold = True
_HEADER_FORMAT.fill.fill_type = style.Fill.FILL_PATTERN_LIGHTGRAY
_HEADER_FORMAT.fill.start_color = style.Color.YELLOW
_HEADER_FORMAT.protection.locked = style.Protection.PROTECTION_PROTECTED


def write_cell_value(value):
    value = flatten_value(value)
    if isinstance(value, (float, int, long, unicode)) or value is None:
        return value
    else:
        return dumps(value, sort_keys=True)


def read_cell_value(value):
    if value == '':
        return None
    try:
        value = loads(value)
    except (TypeError, ValueError):
        pass
    else:
        if isinstance(value, (dict, list, basestring)):
            return flatten_value(value)

    try:
        value = float(value)
    except ValueError:
        return value
    else:
        int_value = int(value)
        if int_value == value:
            return int_value
        else:
            return value


class XLS(object):
    _base = 26
    _A = 65

    @classmethod
    def get_column_name(cls, n):
        n = int(n) + 1
        s = ""
        while n:
            s += chr((n - 1) % cls._base + cls._A)
            n //= cls._base + 1
        return s[::-1]

    @classmethod
    def get_cell_name(cls, column, row):
        """
        This method will transform all good cell coordinates to mf excel AA12 form. Use it for formulas.
        """
        #Convert column number to Excel name.
        return '{}{}'.format(cls.get_column_name(column), row + 1)

    def __init__(self):
        self._sheets = []

    def add_sheet(self, name):
        sheet = Sheet(name)
        self._sheets.append(sheet)
        return sheet

    def get_sheet(self, name):
        try:
            return filter(lambda sheet: sheet.name == name, self._sheets)[0]
        except IndexError:
            raise KeyError('Sheet with name "{}" not found.'.format(name))

    def write(self, stream):
        wb = _Workbook()
        for sheet in self._sheets:
            sheet.write(wb)
        wb.save(stream)

    def read(self, content):
        wb = load_workbook(content)
        for sheet_name in wb.get_sheet_names():
            sheet = self.add_sheet(sheet_name)
            wb_sheet = wb.get_sheet_by_name(sheet_name)
            rows = xrange(wb_sheet.get_highest_row())
            cols = xrange(wb_sheet.get_hightes_col())
            for r in rows:
                row_style = CELL
                row_cells = []
                for c in cols:
                    cell = wb_sheet.cell(row=r, col=c)
                    row_cells.append(cell)
                    cell_style = cell.style
                    if row_style == CELL and cell_style.fill.fill_type == style.Fill.FILL_PATTERN_LIGHTGRAY:
                        row_style = HEADER
                sheet.add_row(row_cells, row_style)


class Sheet(object):
    def __init__(self, name):
        self.name = name
        self.rows = []

    def add_row(self, data, style=CELL):
        row = Row(data, style)
        self.rows.append(row)
        return row

    @property
    def num_rows(self):
        return len(self.rows)

    _styles = {}

    @classmethod
    def get_style(cls, style):
        try:
            return cls._styles[style]
        except KeyError:
            pass

        if style == CELL:
            _style = cls._styles[CELL] = _CELL_FORMAT
        elif style == HEADER:
            _style = cls._styles[HEADER] = _HEADER_FORMAT
        else:
            raise ValueError('Allowed styles are "head", "cell". Got {}.'.format(style))
        return _style

    def write(self, workbook, worksheet=None):
        """ Write data onto workbook. If no worksheet specified, than currently active worksheet used.
        Warning! If all data in worksheet will be erased.

        :param workbook: Workbook instance.
        :param worksheet: Optional worksheet instance.
        """
        if worksheet is None:
            worksheet = workbook.get_active_sheet()
            worksheet.title = self.name

        col_widths = {}
        for row_num, row in enumerate(self.rows):
            style = self.get_style(row.style)
            for cell_num, value in enumerate(row.cells):
                value = write_cell_value(value)
                worksheet.write(row_num, cell_num, value, style)
                col_widths[cell_num] = min(max(col_widths.get(cell_num, 10), len(unicode(value))), 54)

        for num, width in col_widths.iteritems():
            worksheet.col(num).width = int(width * 256 * 1.2)


class Row(object):
    def __init__(self, cells, style=CELL):
        self.cells = cells
        self.style = style